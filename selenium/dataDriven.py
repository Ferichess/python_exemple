from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

from openpyxl import load_workbook

# Load the Excel sheet
workbook = load_workbook(
    filename="C:/feri/lern_node.js/python_exemple/selenium/testdata.xlsx"
)
print(workbook.sheetnames)  # Check if the workbook is loaded correctly
sheet = workbook.active

driver = webdriver.Chrome()
driver.maximize_window()

if sheet is None:
    print("Sheet not loaded properly")
else:
    print("Sheet loaded successfully")
    try:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            username = row[0]
            password = row[1]
            print("checking", row)

            website = "https://www.saucedemo.com/"
            driver.get(website)

            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "user-name"))
            ).send_keys(username)

            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "password"))
            ).send_keys(password)

            WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.ID, "login-button"))
            ).click()

            WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.ID, "react-burger-menu-btn"))
            ).click()

            WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.ID, "logout_sidebar_link"))
            ).click()
    except TimeoutException as e:
        print(f"Timeout occurred: {e}")
    except NoSuchElementException as e:
        print(f"Element not found: {e}")
    finally:
        driver.quit()
